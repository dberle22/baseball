from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from src.pipeline import build_draft_board
from src.standardize import apply_hitter_positions, standardize_projection
from src.yahoo.auth import _is_localhost_callback, _oauth_data_is_present
from src.yahoo.client import get_free_agents


def test_standardize_pitchers_derives_saves_plus_holds() -> None:
    frame = pd.DataFrame(
        {
            "Name": ["Reliever A"],
            "Team": ["AAA"],
            "PlayerId": [1],
            "MLBAMID": [10],
            "ADP": [200],
            "IP": [60],
            "QS": [0],
            "ERA": [3.1],
            "WHIP": [1.05],
            "SO": [70],
            "SV": [20],
            "HLD": [12],
            "GS": [0],
            "G": [62],
            "BB": [18],
        }
    )
    standardized = standardize_projection(frame, "test", "pitchers")
    assert standardized.loc[0, "Saves_plus_Holds"] == 32
    assert standardized.loc[0, "inferred_role"] == "RP"


def test_build_draft_board_creates_outputs(tmp_path: Path) -> None:
    result = build_draft_board(outdir=str(tmp_path), phase="early")
    workbook_path = result["workbook_path"]

    assert workbook_path.exists()
    assert (Path("data/processed") / "hitters_scored.csv").exists()
    assert (Path("data/processed") / "hitters_position_review.csv").exists()
    assert (Path("data/processed") / "pitchers_scored.csv").exists()
    assert (Path("data/processed") / "overall_board.csv").exists()

    with ZipFile(workbook_path) as workbook_zip:
        names = workbook_zip.namelist()
    assert "xl/workbook.xml" in names
    assert any(name.startswith("xl/worksheets/sheet") for name in names)

    workbook = load_workbook(workbook_path, data_only=False)
    assert {"overall_board", "my_team", "targets_by_need"}.issubset(workbook.sheetnames)
    overall_headers = [cell.value for cell in workbook["overall_board"][1]]
    assert "taken" in overall_headers
    assert "dropoff_to_next" in overall_headers
    my_team_headers = [cell.value for cell in workbook["my_team"][3]]
    assert {"position", "filled", "target", "gap"}.issubset(set(my_team_headers))
    assert 'player_pool!$A$2:$A$2000' in workbook["my_team"]["C4"].value
    assert 'player_pool!$B$2:$B$2000' in workbook["my_team"]["C4"].value
    assert '$E4<>"P"' in workbook["my_team"]["O4"].value
    assert 'ARRAY_CONSTRAIN' in workbook["targets_by_need"]["B4"].value
    assert 'SORT(' in workbook["targets_by_need"]["B4"].value


def test_apply_hitter_positions_sets_role_from_lookup() -> None:
    hitters = pd.DataFrame(
        {
            "player_name": ["Jose Ramirez", "Unknown Hitter"],
            "player_name_ascii": ["Jose Ramirez", "Unknown Hitter"],
            "inferred_role": ["BAT", "BAT"],
        }
    )
    positions = pd.DataFrame({"Name": ["Jose Ramirez"], "POS": ["3B"]})

    enriched = apply_hitter_positions(hitters, positions)

    assert enriched.loc[0, "position"] == "3B"
    assert enriched.loc[0, "inferred_role"] == "3B"
    assert pd.isna(enriched.loc[1, "position"])
    assert enriched.loc[1, "inferred_role"] == "BAT"


def test_output_contains_expected_columns(tmp_path: Path) -> None:
    result = build_draft_board(outdir=str(tmp_path), phase="late")
    overall_board = result["overall_board"]
    hitters = result["hitters"]
    hitters_position_review = result["hitters_position_review"]
    queue = result["draft_day_queue"]

    assert {"player_name", "overall_rank", "adp_gap", "tier", "note"}.issubset(overall_board.columns)
    assert "dropoff_to_next" in overall_board.columns
    assert {"position", "inferred_role"}.issubset(hitters.columns)
    assert "dropoff_to_next" in hitters.columns
    assert hitters["inferred_role"].isin(["C", "1B", "2B", "3B", "SS", "OF", "DH", "BAT"]).all()
    assert (hitters_position_review["inferred_role"] == "BAT").all()
    assert {"queue_rank", "queue_score", "note"}.issubset(queue.columns)
    assert queue.iloc[0]["queue_score"] >= queue.iloc[10]["queue_score"]


def test_rebuild_preserves_workbook_inputs(tmp_path: Path) -> None:
    first_result = build_draft_board(outdir=str(tmp_path), phase="early")
    workbook_path = first_result["workbook_path"]

    workbook = load_workbook(workbook_path)
    overall_sheet = workbook["overall_board"]
    overall_sheet["L2"] = "x"
    my_team_sheet = workbook["my_team"]
    my_team_sheet["B4"] = "Shohei Ohtani"
    my_team_sheet["AA4"] = 1.5
    workbook.save(workbook_path)
    workbook.close()

    build_draft_board(outdir=str(tmp_path), phase="late")

    rebuilt = load_workbook(workbook_path, data_only=False)
    assert rebuilt["overall_board"]["L2"].value == "x"
    assert rebuilt["my_team"]["B4"].value == "Shohei Ohtani"
    assert rebuilt["my_team"]["AA4"].value == 1.5
    rebuilt.close()


def test_get_free_agents_dedupes_and_sorts() -> None:
    class StubLeague:
        def free_agents(self, position: str):
            if position == "B":
                return [
                    {"player_id": 2, "name": "Batter Two", "eligible_positions": ["1B"], "percent_owned": 15},
                    {"player_id": 1, "name": "Dual Player", "eligible_positions": ["OF"], "percent_owned": 52},
                ]
            return [
                {"player_id": 1, "name": "Dual Player", "eligible_positions": ["OF", "SP"], "percent_owned": 52},
                {"player_id": 3, "name": "Pitcher Three", "eligible_positions": ["SP"], "percent_owned": 8},
            ]

    from src.yahoo import client as yahoo_client

    original_get_league = yahoo_client._get_league
    try:
        yahoo_client._get_league = lambda sc, league_id=None: StubLeague()
        players = get_free_agents(sc=object(), count=3)
    finally:
        yahoo_client._get_league = original_get_league

    assert [player["player_id"] for player in players] == [1, 2, 3]
    assert players[0]["positions"] == ["OF", "SP"]


def test_normalize_numeric_league_id_from_current_season_candidates() -> None:
    class StubGame:
        def league_ids(self, **kwargs):
            if kwargs.get("seasons"):
                return ["458.l.12345", "458.l.67890"]
            return []

    class StubYfa:
        @staticmethod
        def Game(sc, code):
            assert code == "mlb"
            return StubGame()

    from src.yahoo import client as yahoo_client

    original_import_yfa = yahoo_client._import_yfa
    original_get_league_id = yahoo_client.get_league_id
    try:
        yahoo_client._import_yfa = lambda: StubYfa()
        yahoo_client.get_league_id = lambda: "67890"
        resolved = yahoo_client._normalize_league_id(sc=object())
    finally:
        yahoo_client._import_yfa = original_import_yfa
        yahoo_client.get_league_id = original_get_league_id

    assert resolved == "458.l.67890"


def test_is_localhost_callback_recognizes_local_http_callback() -> None:
    assert _is_localhost_callback("http://localhost:8080/callback") is True
    assert _is_localhost_callback("http://127.0.0.1:8080/callback") is True
    assert _is_localhost_callback("https://localhost:8080/callback") is True
    assert _is_localhost_callback("oob") is False


def test_oauth_data_is_present_requires_access_refresh_and_token_time(tmp_path: Path) -> None:
    token_path = tmp_path / "token.json"
    token_path.write_text('{"access_token": "a", "refresh_token": "b", "token_time": 1}', encoding="utf-8")
    assert _oauth_data_is_present(token_path) is True

    token_path.write_text('{"access_token": "a"}', encoding="utf-8")
    assert _oauth_data_is_present(token_path) is False
