from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


def export_workbook(
    sheets: Dict[str, pd.DataFrame],
    path: Path,
    interactive_context: Dict[str, Any] | None = None,
) -> None:
    preserved_inputs = load_workbook_inputs(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if sheet_name == "overall_board" and preserved_inputs["taken_by_player"]:
                df = df.copy()
                df["taken"] = df["player_name"].map(preserved_inputs["taken_by_player"]).fillna(df["taken"])
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            worksheet = writer.sheets[sheet_name[:31]]
            _format_worksheet(worksheet)
            if sheet_name == "draft_day_queue":
                _add_live_queue_availability(worksheet)

        if interactive_context:
            _add_interactive_sheets(writer, interactive_context, preserved_inputs)


def load_workbook_inputs(path: Path) -> Dict[str, Dict[str, Any]]:
    preserved = {
        "taken_by_player": {},
        "my_team_by_slot": {},
        "manual_need_by_category": {},
    }
    if not path.exists():
        return preserved

    try:
        workbook = load_workbook(path, data_only=False)
    except (BadZipFile, EOFError, OSError):
        return preserved

    if "overall_board" in workbook.sheetnames:
        sheet = workbook["overall_board"]
        headers = [cell.value for cell in sheet[1]]
        try:
            name_col = headers.index("player_name") + 1
            taken_col = headers.index("taken") + 1
        except ValueError:
            name_col = 0
            taken_col = 0
        if name_col and taken_col:
            for row in range(2, sheet.max_row + 1):
                player_name = sheet.cell(row=row, column=name_col).value
                taken_value = sheet.cell(row=row, column=taken_col).value
                if player_name and taken_value not in {None, ""}:
                    preserved["taken_by_player"][str(player_name)] = taken_value

    if "my_team" in workbook.sheetnames:
        sheet = workbook["my_team"]
        for row in range(4, sheet.max_row + 1):
            slot = sheet.cell(row=row, column=1).value
            player_name = sheet.cell(row=row, column=2).value
            category = sheet.cell(row=row, column=22).value
            manual_need = sheet.cell(row=row, column=27).value
            if slot and player_name not in {None, ""}:
                preserved["my_team_by_slot"][str(slot)] = player_name
            if category and manual_need not in {None, ""}:
                preserved["manual_need_by_category"][str(category)] = manual_need

    workbook.close()
    return preserved


def _format_worksheet(worksheet, freeze_panes: str = "A2") -> None:
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 28)


def _add_live_queue_availability(worksheet) -> None:
    taken_col = worksheet.max_column + 1
    available_col = worksheet.max_column + 2
    worksheet.cell(row=1, column=taken_col, value="taken_status").font = Font(bold=True)
    worksheet.cell(row=1, column=available_col, value="available_flag").font = Font(bold=True)

    for row in range(2, worksheet.max_row + 1):
        player_cell = f"C{row}"
        worksheet.cell(
            row=row,
            column=taken_col,
            value=f'=XLOOKUP({player_cell},overall_board!$B$2:$B$2000,overall_board!$L$2:$L$2000,"")',
        )
        worksheet.cell(
            row=row,
            column=available_col,
            value=f'=IF(L{row}="","available","taken")',
        )

    worksheet.auto_filter.ref = worksheet.dimensions


def _add_interactive_sheets(
    writer: pd.ExcelWriter,
    context: Dict[str, Any],
    preserved_inputs: Dict[str, Dict[str, Any]],
) -> None:
    workbook = writer.book
    player_pool = _build_player_pool(context["overall_board"])
    goal_setup = _build_goal_setup(context["hitters"], context["pitchers"], context["league"]["roster"])

    player_pool.to_excel(writer, index=False, sheet_name="player_pool")
    goal_setup.to_excel(writer, index=False, sheet_name="goal_setup")

    player_pool_sheet = writer.sheets["player_pool"]
    goal_setup_sheet = writer.sheets["goal_setup"]
    player_pool_sheet.sheet_state = "hidden"
    goal_setup_sheet.sheet_state = "hidden"
    _format_worksheet(player_pool_sheet)
    _format_worksheet(goal_setup_sheet)

    _build_my_team_sheet(workbook, context["league"]["roster"], preserved_inputs)
    _build_targets_sheet(workbook, len(player_pool), len(context["overall_board"]), context["league"]["roster"])


def _build_player_pool(overall_board: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "player_name",
        "team",
        "inferred_role",
        "player_type",
        "final_score",
        "tier",
        "note",
        "adp",
        "adp_gap",
        "R_z",
        "RBI_z",
        "HR_z",
        "SB_z",
        "AVG_z",
        "OBP_z",
        "QS_z",
        "K_z",
        "Saves_plus_Holds_z",
        "ERA_z",
        "WHIP_z",
        "PA",
        "R",
        "RBI",
        "HR",
        "SB",
        "AVG",
        "OBP",
        "IP",
        "QS",
        "K",
        "Saves_plus_Holds",
        "ERA",
        "WHIP",
    ]
    pool = overall_board.reindex(columns=columns).copy()
    numeric_cols = [col for col in columns if col not in {"player_name", "team", "inferred_role", "player_type", "note"}]
    pool[numeric_cols] = pool[numeric_cols].apply(pd.to_numeric, errors="coerce")
    return pool


def _build_goal_setup(hitters: pd.DataFrame, pitchers: pd.DataFrame, roster: Dict[str, int]) -> pd.DataFrame:
    selected_hitters = _select_goal_hitters(hitters, roster)
    selected_pitchers = _select_goal_pitchers(pitchers, roster)

    goal_rows = [
        ("R", float(selected_hitters["R"].fillna(0).sum())),
        ("RBI", float(selected_hitters["RBI"].fillna(0).sum())),
        ("HR", float(selected_hitters["HR"].fillna(0).sum())),
        ("SB", float(selected_hitters["SB"].fillna(0).sum())),
        ("AVG", _weighted_ratio(selected_hitters, "AVG", "PA")),
        ("OBP", _weighted_ratio(selected_hitters, "OBP", "PA")),
        ("QS", float(selected_pitchers["QS"].fillna(0).sum())),
        ("K", float(selected_pitchers["K"].fillna(0).sum())),
        ("Saves_plus_Holds", float(selected_pitchers["Saves_plus_Holds"].fillna(0).sum())),
        ("ERA", _weighted_ratio(selected_pitchers, "ERA", "IP")),
        ("WHIP", _weighted_ratio(selected_pitchers, "WHIP", "IP")),
    ]
    return pd.DataFrame(goal_rows, columns=["category", "goal_value"])


def _select_goal_hitters(hitters: pd.DataFrame, roster: Dict[str, int]) -> pd.DataFrame:
    ordered = hitters.sort_values("final_score", ascending=False).copy()
    selected_index: List[int] = []
    for position in ["C", "1B", "2B", "3B", "SS"]:
        selected_index.extend(_take_best_available(ordered, selected_index, "inferred_role", position, int(roster.get(position, 0))))
    selected_index.extend(_take_best_available(ordered, selected_index, "inferred_role", "OF", int(roster.get("OF", 0))))
    selected_index.extend(_take_best_available(ordered, selected_index, None, None, int(roster.get("UTIL", 0))))
    return ordered.loc[selected_index].copy()


def _select_goal_pitchers(pitchers: pd.DataFrame, roster: Dict[str, int]) -> pd.DataFrame:
    ordered = pitchers.sort_values("final_score", ascending=False).copy()
    selected_index: List[int] = []
    selected_index.extend(_take_best_available(ordered, selected_index, "inferred_role", "SP", int(roster.get("SP", 0))))
    selected_index.extend(_take_best_available(ordered, selected_index, "inferred_role", "RP", int(roster.get("RP", 0))))
    selected_index.extend(_take_best_available(ordered, selected_index, None, None, int(roster.get("P", 0))))
    return ordered.loc[selected_index].copy()


def _take_best_available(
    df: pd.DataFrame,
    selected_index: Iterable[int],
    filter_col: str | None,
    filter_value: str | None,
    count: int,
) -> List[int]:
    if count <= 0:
        return []
    selected = set(selected_index)
    available = df.loc[~df.index.isin(selected)]
    if filter_col and filter_value is not None:
        available = available.loc[available[filter_col] == filter_value]
    return available.head(count).index.tolist()


def _weighted_ratio(df: pd.DataFrame, ratio_col: str, volume_col: str) -> float:
    values = pd.to_numeric(df.get(ratio_col), errors="coerce").fillna(0.0)
    volume = pd.to_numeric(df.get(volume_col), errors="coerce").fillna(0.0)
    denominator = float(volume.sum())
    if denominator <= 0:
        return 0.0
    return float((values * volume).sum() / denominator)


def _expand_roster_slots(roster: Dict[str, int]) -> List[tuple[str, int]]:
    slots: List[tuple[str, int]] = []
    for slot, count in roster.items():
        if slot == "IL":
            continue
        for index in range(1, int(count) + 1):
            label = slot if count == 1 else f"{slot}{index}"
            include_in_summary = 0 if slot == "BN" else 1
            slots.append((label, include_in_summary))
    return slots


def _build_my_team_sheet(workbook, roster: Dict[str, int], preserved_inputs: Dict[str, Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("my_team")
    slots = _expand_roster_slots(roster)
    end_row = len(slots) + 3

    sheet["A1"] = "Enter your drafted players in column B. Summary updates automatically."
    sheet["A1"].font = Font(bold=True)

    headers = [
        "slot",
        "player_name",
        "team",
        "role",
        "type",
        "final_score",
        "note",
        "R",
        "RBI",
        "HR",
        "SB",
        "AVG",
        "OBP",
        "QS",
        "K",
        "Saves_plus_Holds",
        "ERA",
        "WHIP",
        "PA",
        "IP",
        "include_in_summary",
    ]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col_index, value=header).font = Font(bold=True)

    lookup_ranges = {
        3: "player_pool!$B$2:$B$2000",
        4: "player_pool!$C$2:$C$2000",
        5: "player_pool!$D$2:$D$2000",
        6: "player_pool!$E$2:$E$2000",
        7: "player_pool!$G$2:$G$2000",
        8: "player_pool!$V$2:$V$2000",
        9: "player_pool!$W$2:$W$2000",
        10: "player_pool!$X$2:$X$2000",
        11: "player_pool!$Y$2:$Y$2000",
        12: "player_pool!$Z$2:$Z$2000",
        13: "player_pool!$AA$2:$AA$2000",
        14: "player_pool!$AC$2:$AC$2000",
        15: "player_pool!$AD$2:$AD$2000",
        16: "player_pool!$AE$2:$AE$2000",
        17: "player_pool!$AF$2:$AF$2000",
        18: "player_pool!$AG$2:$AG$2000",
        19: "player_pool!$U$2:$U$2000",
        20: "player_pool!$AB$2:$AB$2000",
    }
    player_pool_lookup = "player_pool!$A$2:$A$2000"

    for row_index, (slot_label, include_in_summary) in enumerate(slots, start=4):
        sheet.cell(row=row_index, column=1, value=slot_label)
        preserved_player = preserved_inputs["my_team_by_slot"].get(slot_label, "")
        sheet.cell(row=row_index, column=2, value=preserved_player)
        sheet.cell(row=row_index, column=21, value=include_in_summary)
        for output_col, return_range in lookup_ranges.items():
            if output_col in {14, 15, 16, 17, 18, 20}:
                formula = f'=IF(OR($B{row_index}="",$E{row_index}<>"P"),"",XLOOKUP($B{row_index},{player_pool_lookup},{return_range},""))'
            else:
                formula = f'=IF($B{row_index}="","",XLOOKUP($B{row_index},{player_pool_lookup},{return_range},""))'
            sheet.cell(row=row_index, column=output_col, value=formula)

    summary_headers = ["category", "current", "goal", "gap", "auto_need", "manual_need", "effective_need"]
    for col_index, header in enumerate(summary_headers, start=22):
        sheet.cell(row=3, column=col_index, value=header).font = Font(bold=True)

    summary_rows = {
        "R": ("=SUMPRODUCT(($U$4:$U$%d=1)*$H$4:$H$%d)" % (end_row, end_row), 2),
        "RBI": ("=SUMPRODUCT(($U$4:$U$%d=1)*$I$4:$I$%d)" % (end_row, end_row), 3),
        "HR": ("=SUMPRODUCT(($U$4:$U$%d=1)*$J$4:$J$%d)" % (end_row, end_row), 4),
        "SB": ("=SUMPRODUCT(($U$4:$U$%d=1)*$K$4:$K$%d)" % (end_row, end_row), 5),
        "AVG": (
            '=IF(SUMPRODUCT(($U$4:$U$%d=1)*$S$4:$S$%d)=0,"",SUMPRODUCT(($U$4:$U$%d=1)*$L$4:$L$%d*$S$4:$S$%d)/SUMPRODUCT(($U$4:$U$%d=1)*$S$4:$S$%d))'
            % (end_row, end_row, end_row, end_row, end_row, end_row, end_row),
            6,
        ),
        "OBP": (
            '=IF(SUMPRODUCT(($U$4:$U$%d=1)*$S$4:$S$%d)=0,"",SUMPRODUCT(($U$4:$U$%d=1)*$M$4:$M$%d*$S$4:$S$%d)/SUMPRODUCT(($U$4:$U$%d=1)*$S$4:$S$%d))'
            % (end_row, end_row, end_row, end_row, end_row, end_row, end_row),
            7,
        ),
        "QS": ("=SUMPRODUCT(($U$4:$U$%d=1)*$N$4:$N$%d)" % (end_row, end_row), 8),
        "K": ("=SUMPRODUCT(($U$4:$U$%d=1)*$O$4:$O$%d)" % (end_row, end_row), 9),
        "Saves_plus_Holds": ("=SUMPRODUCT(($U$4:$U$%d=1)*$P$4:$P$%d)" % (end_row, end_row), 10),
        "ERA": (
            '=IF(SUMPRODUCT(($U$4:$U$%d=1)*$T$4:$T$%d)=0,"",SUMPRODUCT(($U$4:$U$%d=1)*$Q$4:$Q$%d*$T$4:$T$%d)/SUMPRODUCT(($U$4:$U$%d=1)*$T$4:$T$%d))'
            % (end_row, end_row, end_row, end_row, end_row, end_row, end_row),
            11,
        ),
        "WHIP": (
            '=IF(SUMPRODUCT(($U$4:$U$%d=1)*$T$4:$T$%d)=0,"",SUMPRODUCT(($U$4:$U$%d=1)*$R$4:$R$%d*$T$4:$T$%d)/SUMPRODUCT(($U$4:$U$%d=1)*$T$4:$T$%d))'
            % (end_row, end_row, end_row, end_row, end_row, end_row, end_row),
            12,
        ),
    }

    for row_index, (category, (current_formula, goal_row)) in enumerate(summary_rows.items(), start=4):
        sheet.cell(row=row_index, column=22, value=category)
        sheet.cell(row=row_index, column=23, value=current_formula)
        sheet.cell(row=row_index, column=24, value=f'=XLOOKUP(V{row_index},goal_setup!$A$2:$A$12,goal_setup!$B$2:$B$12,"")')
        if category in {"ERA", "WHIP"}:
            gap_formula = f'=IF(OR(W{row_index}="",X{row_index}=""),"",W{row_index}-X{row_index})'
        else:
            gap_formula = f'=IF(OR(W{row_index}="",X{row_index}=""),"",X{row_index}-W{row_index})'
        sheet.cell(row=row_index, column=25, value=gap_formula)
        sheet.cell(row=row_index, column=26, value=f'=IF(OR(Y{row_index}="",X{row_index}=0),"",MAX(MIN((Y{row_index}/ABS(X{row_index}))*4,2),-2))')
        sheet.cell(
            row=row_index,
            column=27,
            value=preserved_inputs["manual_need_by_category"].get(category, ""),
        )
        sheet.cell(row=row_index, column=28, value=f'=IF(AA{row_index}="",Z{row_index},AA{row_index})')

    position_headers = ["position", "filled", "target", "gap"]
    for col_index, header in enumerate(position_headers, start=30):
        sheet.cell(row=3, column=col_index, value=header).font = Font(bold=True)

    position_targets = {
        "C": int(roster.get("C", 0)),
        "1B": int(roster.get("1B", 0)),
        "2B": int(roster.get("2B", 0)),
        "3B": int(roster.get("3B", 0)),
        "SS": int(roster.get("SS", 0)),
        "OF": int(roster.get("OF", 0)),
        "SP": int(roster.get("SP", 0)),
        "RP": int(roster.get("RP", 0)),
        "P": int(roster.get("P", 0)),
    }
    for row_index, (position, target) in enumerate(position_targets.items(), start=4):
        sheet.cell(row=row_index, column=30, value=position)
        if position == "P":
            filled_formula = f'=COUNTIF($D$4:$D${end_row},"SP")+COUNTIF($D$4:$D${end_row},"RP")+COUNTIF($D$4:$D${end_row},"P")'
        else:
            filled_formula = f'=COUNTIF($D$4:$D${end_row},AD{row_index})'
        sheet.cell(row=row_index, column=31, value=filled_formula)
        sheet.cell(row=row_index, column=32, value=target)
        sheet.cell(row=row_index, column=33, value=f'=AF{row_index}-AE{row_index}')

    for col in ["S", "T", "U"]:
        sheet.column_dimensions[col].hidden = True
    _format_worksheet(sheet, freeze_panes="A4")


def _build_targets_sheet(workbook, player_pool_rows: int, overall_board_rows: int, roster: Dict[str, int]) -> None:
    sheet = workbook.create_sheet("targets_by_need")
    roster_end_row = len(_expand_roster_slots(roster)) + 3

    sheet["A1"] = "Best available targets based on final score, your current roster, and category needs from my_team."
    sheet["A1"].font = Font(bold=True)
    headers = ["target_rank", "player_name", "team", "role", "type", "target_score", "final_score", "adp_gap", "note"]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col_index, value=header).font = Font(bold=True)

    player_pool_end = player_pool_rows + 1
    overall_end = overall_board_rows + 1
    formula = f"""=LET(
names,player_pool!$A$2:$A${player_pool_end},
teams,player_pool!$B$2:$B${player_pool_end},
roles,player_pool!$C$2:$C${player_pool_end},
types,player_pool!$D$2:$D${player_pool_end},
finals,player_pool!$E$2:$E${player_pool_end},
notes,player_pool!$G$2:$G${player_pool_end},
adpgaps,player_pool!$I$2:$I${player_pool_end},
rz,player_pool!$J$2:$J${player_pool_end},
rbiz,player_pool!$K$2:$K${player_pool_end},
hrz,player_pool!$L$2:$L${player_pool_end},
sbz,player_pool!$M$2:$M${player_pool_end},
avgz,player_pool!$N$2:$N${player_pool_end},
obpz,player_pool!$O$2:$O${player_pool_end},
qsz,player_pool!$P$2:$P${player_pool_end},
kz,player_pool!$Q$2:$Q${player_pool_end},
svhz,player_pool!$R$2:$R${player_pool_end},
eraz,player_pool!$S$2:$S${player_pool_end},
whipz,player_pool!$T$2:$T${player_pool_end},
takenflags,XLOOKUP(names,overall_board!$B$2:$B${overall_end},overall_board!$L$2:$L${overall_end},""),
needscore,ARRAYFORMULA(finals+IF(types="H",rz*my_team!$AB$4+rbiz*my_team!$AB$5+hrz*my_team!$AB$6+sbz*my_team!$AB$7+avgz*my_team!$AB$8+obpz*my_team!$AB$9,qsz*my_team!$AB$10+kz*my_team!$AB$11+svhz*my_team!$AB$12+eraz*my_team!$AB$13+whipz*my_team!$AB$14)),
available,ARRAYFORMULA((((takenflags="")+(takenflags=0))>0)*(COUNTIF(my_team!$B$4:$B${roster_end_row},names)=0)),
filtered,FILTER({{names,teams,roles,types,needscore,finals,adpgaps,notes}},available),
sorted,SORT(filtered,5,FALSE,6,FALSE),
ARRAY_CONSTRAIN(sorted,75,8)
)"""
    sheet["B4"] = formula
    sheet["A4"] = '=ARRAYFORMULA(IF(B4:B78="","",ROW(B4:B78)-3))'
    _format_worksheet(sheet, freeze_panes="A4")
